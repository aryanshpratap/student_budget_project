"""
Build Student_Budget_Analysis.xlsx
Sheets: Transactions, Budget, Monthly Summary, Budget vs Actual, Dashboard
Uses live Excel formulas (SUMIFS, etc.) - not hardcoded Python-calculated values.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import LineChart, PieChart, BarChart, Reference

CLEAN_PATH = "data/cleaned_finance_transactions.csv"
OUT_PATH = "Student_Budget_Analysis.xlsx"

CURRENCY_FMT = '"Rs."#,##0;("Rs."#,##0);"-"'
PCT_FMT = '0.0%'

NAVY = "1F3864"
HEADER_FILL = PatternFill("solid", start_color=NAVY)
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
INPUT_FONT = Font(name="Arial", color="0000FF")
CALC_FONT = Font(name="Arial", color="000000")
BOLD_FONT = Font(name="Arial", bold=True)
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

month_names = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]

categories = [
    ("Food & Snacks", 2500),
    ("Transportation", 600),
    ("Mobile & Internet", 350),
    ("Education", 400),
    ("Entertainment", 500),
    ("Subscriptions", 250),
    ("Personal Care", 200),
    ("Clothing", 400),
    ("Savings", 1500),
    ("Miscellaneous", 200),
]

df = pd.read_csv(CLEAN_PATH, parse_dates=["Date"])
df = df[["Date", "Type", "Category", "Description", "Amount", "Payment_Mode"]]

wb = Workbook()

# ============================================================
# SHEET 1: Transactions
# ============================================================
ws = wb.active
ws.title = "Transactions"
headers = ["Date", "Type", "Category", "Description", "Amount", "Payment Mode", "Month", "Month Name"]
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center")

n_rows = len(df)
for i, row in enumerate(df.itertuples(index=False), start=2):
    ws.cell(row=i, column=1, value=row.Date.to_pydatetime()).number_format = "dd-mmm-yyyy"
    ws.cell(row=i, column=2, value=row.Type)
    ws.cell(row=i, column=3, value=row.Category)
    ws.cell(row=i, column=4, value=row.Description)
    amt_cell = ws.cell(row=i, column=5, value=float(row.Amount))
    amt_cell.number_format = CURRENCY_FMT
    ws.cell(row=i, column=6, value=row.Payment_Mode)
    ws.cell(row=i, column=7, value=f"=MONTH(A{i})")
    ws.cell(row=i, column=8, value=f'=TEXT(A{i},"mmm")')

col_widths = [12, 9, 16, 26, 11, 13, 7, 10]
for c, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(c)].width = w

LAST_TX_ROW = n_rows + 1  # last row number with data

# ============================================================
# SHEET 2: Budget
# ============================================================
ws2 = wb.create_sheet("Budget")
headers2 = ["Category", "Monthly Budget", "Annual Budget"]
for c, h in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=c, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center")

for i, (cat, monthly_budget) in enumerate(categories, start=2):
    ws2.cell(row=i, column=1, value=cat)
    b_cell = ws2.cell(row=i, column=2, value=monthly_budget)
    b_cell.font = INPUT_FONT
    b_cell.number_format = CURRENCY_FMT
    a_cell = ws2.cell(row=i, column=3, value=f"=B{i}*12")
    a_cell.number_format = CURRENCY_FMT

total_row = len(categories) + 2
ws2.cell(row=total_row, column=1, value="Total").font = BOLD_FONT
tc = ws2.cell(row=total_row, column=2, value=f"=SUM(B2:B{total_row-1})")
tc.font = BOLD_FONT
tc.number_format = CURRENCY_FMT
tc2 = ws2.cell(row=total_row, column=3, value=f"=SUM(C2:C{total_row-1})")
tc2.font = BOLD_FONT
tc2.number_format = CURRENCY_FMT

for c, w in enumerate([18, 15, 15], 1):
    ws2.column_dimensions[get_column_letter(c)].width = w

# ============================================================
# SHEET 3: Monthly Summary
# ============================================================
ws3 = wb.create_sheet("Monthly Summary")
headers3 = ["Month", "Month #", "Income", "Expense", "Net Savings", "Savings Rate"]
for c, h in enumerate(headers3, 1):
    cell = ws3.cell(row=1, column=c, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center")

for i, m_name in enumerate(month_names, start=2):
    m_num = i - 1
    ws3.cell(row=i, column=1, value=m_name)
    ws3.cell(row=i, column=2, value=m_num)
    inc = ws3.cell(row=i, column=3,
        value=f'=SUMIFS(Transactions!E2:E{LAST_TX_ROW},Transactions!B2:B{LAST_TX_ROW},"Income",Transactions!G2:G{LAST_TX_ROW},B{i})')
    inc.number_format = CURRENCY_FMT
    exp = ws3.cell(row=i, column=4,
        value=f'=SUMIFS(Transactions!E2:E{LAST_TX_ROW},Transactions!B2:B{LAST_TX_ROW},"Expense",Transactions!G2:G{LAST_TX_ROW},B{i})')
    exp.number_format = CURRENCY_FMT
    net = ws3.cell(row=i, column=5, value=f"=C{i}-D{i}")
    net.number_format = CURRENCY_FMT
    rate = ws3.cell(row=i, column=6, value=f"=IF(C{i}=0,0,E{i}/C{i})")
    rate.number_format = PCT_FMT

ms_total_row = len(month_names) + 2
ws3.cell(row=ms_total_row, column=1, value="Total / Avg").font = BOLD_FONT
for col, letter in [(3, "C"), (4, "D"), (5, "E")]:
    cell = ws3.cell(row=ms_total_row, column=col, value=f"=SUM({letter}2:{letter}{ms_total_row-1})")
    cell.font = BOLD_FONT
    cell.number_format = CURRENCY_FMT
avg_cell = ws3.cell(row=ms_total_row, column=6, value=f"=AVERAGE(F2:F{ms_total_row-1})")
avg_cell.font = BOLD_FONT
avg_cell.number_format = PCT_FMT

for c, w in enumerate([12, 9, 13, 13, 13, 13], 1):
    ws3.column_dimensions[get_column_letter(c)].width = w

# ============================================================
# SHEET 4: Budget vs Actual
# ============================================================
ws4 = wb.create_sheet("Budget vs Actual")
headers4 = ["Category", "Annual Budget", "Actual Spent", "Variance", "% of Budget Used"]
for c, h in enumerate(headers4, 1):
    cell = ws4.cell(row=1, column=c, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center")

for i, (cat, _) in enumerate(categories, start=2):
    ws4.cell(row=i, column=1, value=cat)
    budget_cell = ws4.cell(row=i, column=2, value=f"=Budget!C{i}")
    budget_cell.number_format = CURRENCY_FMT
    actual_cell = ws4.cell(row=i, column=3,
        value=f'=SUMIFS(Transactions!E2:E{LAST_TX_ROW},Transactions!B2:B{LAST_TX_ROW},"Expense",Transactions!C2:C{LAST_TX_ROW},A{i})')
    actual_cell.number_format = CURRENCY_FMT
    var_cell = ws4.cell(row=i, column=4, value=f"=B{i}-C{i}")
    var_cell.number_format = CURRENCY_FMT
    pct_cell = ws4.cell(row=i, column=5, value=f"=IF(B{i}=0,0,C{i}/B{i})")
    pct_cell.number_format = PCT_FMT

bva_total_row = len(categories) + 2
ws4.cell(row=bva_total_row, column=1, value="Total").font = BOLD_FONT
for col, letter in [(2, "B"), (3, "C"), (4, "D")]:
    cell = ws4.cell(row=bva_total_row, column=col, value=f"=SUM({letter}2:{letter}{bva_total_row-1})")
    cell.font = BOLD_FONT
    cell.number_format = CURRENCY_FMT
pct_total = ws4.cell(row=bva_total_row, column=5, value=f"=IF(B{bva_total_row}=0,0,C{bva_total_row}/B{bva_total_row})")
pct_total.font = BOLD_FONT
pct_total.number_format = PCT_FMT

# Conditional formatting on % of Budget Used
red_fill = PatternFill("solid", start_color="F8D7DA")
green_fill = PatternFill("solid", start_color="D4EDDA")
yellow_fill = PatternFill("solid", start_color="FFF3CD")
pct_range = f"E2:E{bva_total_row-1}"
ws4.conditional_formatting.add(pct_range, CellIsRule(operator="greaterThan", formula=["1"], fill=red_fill))
ws4.conditional_formatting.add(pct_range, CellIsRule(operator="between", formula=["0.9", "1"], fill=yellow_fill))
ws4.conditional_formatting.add(pct_range, CellIsRule(operator="lessThan", formula=["0.9"], fill=green_fill))

for c, w in enumerate([18, 14, 14, 12, 16], 1):
    ws4.column_dimensions[get_column_letter(c)].width = w

# ============================================================
# SHEET 5: Dashboard
# ============================================================
ws5 = wb.create_sheet("Dashboard", 0)  # make it the first sheet

ws5["A1"] = "STUDENT BUDGET DASHBOARD - 2025"
ws5["A1"].font = Font(name="Arial", bold=True, size=16, color=NAVY)
ws5.merge_cells("A1:F1")

kpi_labels = ["Total Income", "Total Expense", "Net Savings", "Avg Savings Rate"]
kpi_formulas = [
    f"='Monthly Summary'!C{ms_total_row}",
    f"='Monthly Summary'!D{ms_total_row}",
    f"='Monthly Summary'!E{ms_total_row}",
    f"='Monthly Summary'!F{ms_total_row}",
]
kpi_formats = [CURRENCY_FMT, CURRENCY_FMT, CURRENCY_FMT, PCT_FMT]

for i, (label, formula, fmt) in enumerate(zip(kpi_labels, kpi_formulas, kpi_formats)):
    col = 1 + i * 2  # A, C, E, G... use 2-col spacing: place at columns A,C,E,G -> 1,3,5,7
    lcell = ws5.cell(row=3, column=col, value=label)
    lcell.font = Font(name="Arial", bold=True, size=10)
    vcell = ws5.cell(row=4, column=col, value=formula)
    vcell.font = Font(name="Arial", bold=True, size=14, color=NAVY)
    vcell.number_format = fmt

for c in range(1, 8):
    ws5.column_dimensions[get_column_letter(c)].width = 14

# Chart 1: Monthly Income/Expense/Savings line chart
line = LineChart()
line.title = "Monthly Income, Expense & Net Savings"
line.style = 10
line.y_axis.title = "Amount (Rs.)"
data = Reference(ws3, min_col=3, max_col=5, min_row=1, max_row=ms_total_row-1)
cats = Reference(ws3, min_col=1, min_row=2, max_row=ms_total_row-1)
line.add_data(data, titles_from_data=True)
line.set_categories(cats)
line.height = 8
line.width = 16
ws5.add_chart(line, "A6")

# Chart 2: Category breakdown pie chart (Actual Spent)
pie = PieChart()
pie.title = "Spending by Category (Actual)"
pie_data = Reference(ws4, min_col=3, min_row=1, max_row=bva_total_row-1)
pie_cats = Reference(ws4, min_col=1, min_row=2, max_row=bva_total_row-1)
pie.add_data(pie_data, titles_from_data=True)
pie.set_categories(pie_cats)
pie.height = 8
pie.width = 12
ws5.add_chart(pie, "A22")

# Chart 3: Budget vs Actual bar chart
bar = BarChart()
bar.type = "col"
bar.title = "Budget vs Actual Spending by Category"
bar.y_axis.title = "Amount (Rs.)"
bar_data = Reference(ws4, min_col=2, max_col=3, min_row=1, max_row=bva_total_row-1)
bar_cats = Reference(ws4, min_col=1, min_row=2, max_row=bva_total_row-1)
bar.add_data(bar_data, titles_from_data=True)
bar.set_categories(bar_cats)
bar.height = 8
bar.width = 16
ws5.add_chart(bar, "H6")

wb.save(OUT_PATH)
print(f"Saved workbook to {OUT_PATH}")
print(f"Transactions: {n_rows} rows, last row = {LAST_TX_ROW}")
